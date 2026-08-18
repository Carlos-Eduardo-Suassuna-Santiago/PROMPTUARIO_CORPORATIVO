"""
XLSX workbook builder — single and multi-sheet.
Standalone module with no external dependencies beyond openpyxl.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


def build_xlsx(data: list[dict[str, Any]], sheet_name: str = "Report") -> bytes:
    """Build a single-sheet XLSX workbook from a list of dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel max sheet name length

    if not data:
        ws.cell(row=1, column=1, value="No data available")
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    headers = list(data[0].keys())
    _apply_header_style(ws, headers)
    _apply_data_rows(ws, data, headers)
    _auto_fit_columns(ws, data, headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_multi_sheet_xlsx(sheets_data: list[dict[str, Any]]) -> bytes:
    """Build an XLSX workbook with multiple sheets.
    Each element in sheets_data: {"sheet_name": str, "data": list[dict]}"""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets_data:
        sheet_name = sheet_def.get("sheet_name", "Sheet")[:31]
        data = sheet_def.get("data", [])

        ws = wb.create_sheet(title=sheet_name)

        if not data:
            ws.cell(row=1, column=1, value="No data available")
            continue

        headers = list(data[0].keys())
        _apply_header_style(ws, headers)
        _apply_data_rows(ws, data, headers)
        _auto_fit_columns(ws, data, headers)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Private helpers ─────────────────────────────────────────────────────────

def _apply_header_style(ws, headers: list[str]) -> None:
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _apply_data_rows(ws, data: list[dict], headers: list[str]) -> None:
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.border = thin_border


def _auto_fit_columns(ws, data: list[dict], headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, min(len(data) + 2, 100)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 60)